import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

def save_to_excel(client_name, supplier, invoice_number, invoice_date, products, note=""):
    client_folder = os.path.join("clients_data", client_name)
    os.makedirs(client_folder, exist_ok=True)
    month_key = datetime.now().strftime("%Y_%m")
    file_name = os.path.join(client_folder, f"{month_key}.xlsx")

    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "اسم المورد", "رقم الفاتورة", "تاريخ الفاتورة",
            "اسم المنتج", "الفئة", "الصنف المقابل", "الكمية",
            "الوحدة", "سعر الوحدة", "السعر قبل الخصم", "الخصم",
            "السعر بعد الخصم", "نسبة الضريبة", "قيمة الضريبة", "السعر النهائي",
            "لينك الفاتورة", "ملاحظات"
        ])
        wb.save(file_name)

    wb = load_workbook(file_name)
    ws = wb.active

    if not products:
        ws.append([
            supplier, invoice_number, invoice_date,
            "❌ فاتورة غير واضحة", "", "", "", "", "", "", "", "", "", "", "", note or "لم يتم التعرف على المنتجات"
        ])
    else:
        for p in products:
            image_link = p.get("image_url", "")
            ws.append([
                supplier, invoice_number, invoice_date,
                p.get("product_name"), p.get("category", ""),
                p.get("matched_inventory", ""), p.get("quantity"),
                p.get("unit_of_measure", ""),
                p.get("unit_price"), p.get("total_before_discount"),
                p.get("discount_amount", 0.0),
                p.get("total_after_discount", p.get("total_before_discount")),
                p.get("vat_percentage", 0.0),
                p.get("vat_amount", 0.0),
                p.get("item_total", p.get("final_total_per_product", 0.0)),
                f'=HYPERLINK("{image_link}", "عرض الصورة")' if image_link else "",
                note
            ])

    wb.save(file_name)
