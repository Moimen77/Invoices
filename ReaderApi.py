from fastapi import FastAPI, UploadFile, Form
import google.generativeai as genai
from PIL import Image
import json, os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import mysql.connector
from rapidfuzz import fuzz

app = FastAPI()

# 🔑 API Key
genai.configure(api_key="AIzaSyBDhIS_HQauE6BLCwD0qoa2MWGJHTLDIwk")

# ✅ الموديل اللي بيدعم الصور
model = genai.GenerativeModel("gemini-2.5-flash-preview-05-20")


# 🧹 دالة تنظيف رد Gemini
def clean_response(raw_text: str) -> str:
    text = raw_text.strip()
    if text.startswith("```"):
        text = text.strip("`")
        if text.lower().startswith("json"):
            text = text[4:].strip()
    return text


# 📌 إحضار المخزون
def get_inventory_from_db(client_id: int):
    conn = mysql.connector.connect(
        host="localhost", user="root", password="", database="raqiib_invoices"
    )
    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        "SELECT item_id, product_name, unit_price FROM inventory WHERE client_id = %s",
        (client_id,),
    )
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows


# 📌 إحضار الكلمات المفتاحية للعميل
def get_keywords_from_db(client_id: int):
    conn = mysql.connector.connect(
        host="localhost", user="root", password="", database="raqiib_invoices"
    )
    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT k.keyword, k.lang, i.item_id, i.product_name, i.unit_price
        FROM keywords k
        JOIN inventory i ON k.item_id = i.item_id
        WHERE i.client_id = %s
        """,
        (client_id,),
    )
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows


# 📌 المطابقة باستخدام الكلمات المفتاحية + السعر + تحسينات
def match_products_with_inventory(products, keywords_list, threshold=65, price_tolerance=10):
    matches = []

    for p in products:
        name_ar = str(p.get("product_name", "")).lower().strip()
        name_en = str(p.get("alt_name", "")).lower().strip()

        best_match = None
        best_score = 0
        best_item = None

        for kw in keywords_list:
            keyword = str(kw["keyword"]).lower().strip()
            inv_name = str(kw["product_name"]).lower()

            # ✅ تطابق كامل (أولوية قصوى)
            if name_ar == keyword or name_en == keyword:
                score = 120
            else:
                # ✅ لو الكلمة قصيرة جدا (<=3) قارن تطابق كامل فقط
                if len(name_ar) <= 3:
                    score = 100 if name_ar == keyword else 0
                else:
                    score = max(
                        fuzz.partial_ratio(name_ar, keyword),
                        fuzz.partial_ratio(name_en, keyword),
                    )

            # ✅ لو السعر قريب ضيف bonus
            if abs(float(p.get("unit_price", 0)) - float(kw.get("unit_price", 0))) <= price_tolerance:
                score += 10

            if score > best_score:
                best_score = score
                best_match = inv_name
                best_item = kw

        if best_score >= threshold:
            matches.append({
                "product_name": p.get("product_name"),
                "alt_name": p.get("alt_name", ""),
                "matched_inventory": best_match,
                "score": best_score,
                "unit_price_invoice": p.get("unit_price", 0),
                "unit_price_db": best_item.get("unit_price", 0) if best_item else None
            })
        else:
            matches.append({
                "product_name": p.get("product_name"),
                "alt_name": p.get("alt_name", ""),
                "matched_inventory": "",
                "score": best_score
            })

    return matches


# 📌 حفظ في Excel
# 📌 حفظ في Excel (معدل)
def save_to_excel(client_name, supplier, invoice_number, invoice_date, products):
    client_folder = os.path.join("clients_data", client_name)
    os.makedirs(client_folder, exist_ok=True)

    month_key = datetime.now().strftime("%Y_%m")
    file_name = os.path.join(client_folder, f"{month_key}.xlsx")

    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Supplier", "Invoice Number", "Invoice Date",
            "Product","Category", "Matched Inventory", "Quantity",
            "Unit Price", "Total Before Discount", "Discount",
            "Total After Discount", "VAT %", "VAT Amount", "Final Total"
        ])
        wb.save(file_name)

    wb = load_workbook(file_name)
    ws = wb.active

    if not products:  # ❌ لو مفيش منتجات
        ws.append([
            supplier,
            invoice_number,
            invoice_date,
            "❌ لم يتم استخراج منتجات",
            "", "", "", "", "", "", "", "", "", ""
        ])
    else:
        for p in products:
            ws.append([
                supplier,
                invoice_number,
                invoice_date,
                p.get("product_name"),
                p.get("category", ""),
                p.get("matched_inventory", ""),
                p.get("quantity"),
                p.get("unit_price"),
                p.get("total_before_discount"),
                p.get("discount_amount", 0.0),
                p.get("total_after_discount", p.get("total_before_discount")),
                p.get("vat_percentage", 0.0),
                p.get("vat_amount", 0.0),
                p.get("item_total", p.get("final_total_per_product", 0.0))
            ])

    wb.save(file_name)



@app.post("/upload_invoice/")
async def upload_invoice(file: UploadFile, client_id: int = Form(...), client_name: str = Form(...)):
    img = Image.open(file.file)

    # 📜 Prompt فيه category + alt_name
    prompt =  """
حلل لي فاتورة المرفقة. استخرج كل التفاصيل الممكنة في صيغة JSON، مع التأكد من مطابقة جميع الحسابات.

📌 مهم جداً:
- لو المنتج مكتوب بلغتين (عربي + إنجليزي)، خُد العربي كـ "product_name".
- احفظ الإنجليزي (لو موجود) في حقل إضافي اسمه "alt_name".
- الأولوية دائماً للاسم العربي في المطابقة والحفظ.

بيانات الفاتورة: رقم الفاتورة، تاريخ الإصدار، بيانات المورد والعميل.
تفاصيل المنتجات: اسم المنتج (بالعربي إن وجد)، الاسم البديل بالإنجليزي (alt_name)، الكمية، وحدة القياس، سعر الوحدة، الإجمالي قبل الخصم، مبلغ الخصم (إذا وُجد)، الإجمالي بعد الخصم، نسبة الضريبة، مبلغ الضريبة، والإجمالي النهائي للمنتج.
➡️ صنّف كل منتج في خانة category حسب طبيعته (مثال: مواد مخزنية، لحوم، مشروبات، مياه، مخبوزات، منتجات ألبان، مصروف تشغيل، صيانة ,معدات، مواد نظافة، مصارف شحن، غاز، بنزين، مستشفى).

ملخص الفاتورة: الإجمالي الإجمالي، إجمالي الخصم، الإجمالي الصافي قبل الضريبة، إجمالي الضريبة، والإجمالي النهائي للفاتورة.

❌ لا تضع أي شرح أو نص خارجي أو علامات ```json.
✅ أرجع فقط JSON مطابق للشكل التالي وبنفس أسماء المفاتيح بالضبط:

{
  "invoice_details": {
    "invoice_number": "string",
    "issue_date": "string",
    "issue_time": "string"
  },
  "supplier_details": {
    "name": "string",
    "vat_id": "string",
    "address": "string"
  },
  "client_details": {
    "name": "string",
    "account_number": "string",
    "type": "string",
    "address": "string",
    "vat_id": "string"
  },
  "product_details": [
    {
      "product_name": "string",   // بالعربي إن وجد
      "alt_name": "string",       // بالإنجليزي لو موجود
      "quantity": 0,
      "unit_of_measure": "string",
      "unit_price": 0.0,
      "total_before_discount": 0.0,
      "discount_amount": 0.0,
      "total_after_discount": 0.0,
      "vat_percentage": 0.0,
      "vat_amount": 0.0,
      "final_total_per_product": 0.0,
      "category": "string"
    }
  ],
  "invoice_summary": {
    "total_amount_before_vat": 0.0,
    "total_discount_amount": 0.0,
    "net_amount_before_vat": 0.0,
    "total_vat_amount": 0.0,
    "vat_percentage_summary": 0.0,
    "final_invoice_total": 0.0,
    "amount_in_words_ar": "string"
  }
}
"""

    response = model.generate_content([prompt, img])
    raw_text = clean_response(response.text)

    try:
        data = json.loads(raw_text)
    except json.JSONDecodeError:
        return {"error": "الرد مش JSON صالح", "raw": raw_text}

    invoice = data.get("invoice_details", {})
    supplier = data.get("supplier_details", {}).get("name", "")
    products = data.get("product_details", [])

    # 📌 إحضار الكلمات المفتاحية للعميل
    keywords = get_keywords_from_db(client_id)

    # 📌 مطابقة المنتجات مع الكلمات المفتاحية
    matches = match_products_with_inventory(products, keywords)

    # 📌 دمج النتيجة مع المنتجات
    for p in products:
        match = next((m for m in matches if m["product_name"] == p["product_name"]), None)
        if match:
            p["matched_inventory"] = match["matched_inventory"]
        else:
            p["matched_inventory"] = ""

    # 📝 حفظ في Excel
    save_to_excel(
        client_name=client_name,
        supplier=supplier,
        invoice_number=invoice.get("invoice_number", ""),
        invoice_date=invoice.get("issue_date", ""),
        products=products
    )

    return {
    "status": "تم الحفظ بنجاح",
    "rows_added": len(products),
    "warning": "⚠️ لم يتم استخراج منتجات من الفاتورة" if not products else "",
    "parsed_json": data,
    "matches": matches
}

